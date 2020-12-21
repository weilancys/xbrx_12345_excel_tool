import xlrd
import openpyxl
import datetime
import os
from jinja2 import Environment, PackageLoader, select_autoescape
from .utils import str_to_dt, get_now_report_str, dt_to_datetime_str, make_config_dirs


class ZhengFuBiao(object):
    def __init__(self, source_file):
        self.source_file = source_file
        self.rows = []
        self.DATA_STARTS_AT_ROW = 3 # by excel file
        self.COL_RECEIVE_TIME = 2 # by excel file
        self.COL_ID = 1 # by excel file

        if source_file.endswith(".xls"):
            workbook = xlrd.open_workbook(source_file)
            sheet = workbook.sheet_by_index(0)

            self.row_count = sheet.nrows
            self.col_count = sheet.ncols

            for row_index in range(self.DATA_STARTS_AT_ROW - 1, self.row_count):
                row = []
                for col_index in range(self.col_count):
                    value = str(sheet.cell_value(row_index, col_index)).strip()
                    row.append(value)
                self.rows.append(row)

        elif source_file.endswith(".xlsx"):
            workbook = openpyxl.load_workbook(source_file)
            sheet = workbook.active
            for row in sheet.iter_rows(min_row=self.DATA_STARTS_AT_ROW, values_only=True):
                self.rows.append(row)
            
            self.row_count = len(self.rows)
            self.col_count = len(self.rows[0])


    def get_all_rows(self):
        return self.rows


    def get_rows_by_time(self, start_datetime, end_datetime):
        rows = []
        for row in self.rows:
            receive_time_dt = str_to_dt(row[self.COL_RECEIVE_TIME - 1].strip())
            if receive_time_dt >= start_datetime and receive_time_dt <= end_datetime:
                rows.append(row)
        return rows

    
    def get_rows_by_ids(self, ids):
        rows = []
        for row in self.rows:
            if row[self.COL_ID - 1] in ids:
                rows.append(row)
        return rows

    
    def get_earliest_row(self):
        earliest_row = None
        for row in self.rows:
            if earliest_row is None:
                earliest_row = row
            elif str_to_dt(row[self.COL_RECEIVE_TIME - 1].strip()) < str_to_dt(earliest_row[self.COL_RECEIVE_TIME - 1].strip()):
                earliest_row = row
        return earliest_row


    def get_latest_row(self):
        latest_row = None
        for row in self.rows:
            if latest_row is None:
                latest_row = row
            elif str_to_dt(row[self.COL_RECEIVE_TIME - 1].strip()) > str_to_dt(latest_row[self.COL_RECEIVE_TIME - 1].strip()):
                latest_row = row
        return latest_row

    
    def get_all_ids(self, include_duplicates=True):
        ids = []
        for row in self.rows:
            _id = row[0]
            if include_duplicates == False and _id in ids:
                continue
            ids.append(_id)
        return ids


class SangaoBiao(object):
    def __init__(self, source_file):
        self.source_file = source_file
        workbook = openpyxl.load_workbook(source_file)
        sheet_1 = workbook.active

        self.DATA_STARTS_AT_ROW = 3 # by excel file
        self.COL_12345_ID = 2 # by excel file

        self.rows = []
        for row in sheet_1.iter_rows(min_row=self.DATA_STARTS_AT_ROW, values_only=True):
            self.rows.append(row)

        if self.rows == []:
            raise ValueError("empty excel file")
        
        self.row_count = len(self.rows)
        self.col_count = len(self.rows[0])
    

    def get_all_12345_ids(self, include_duplicates=True):
        ids = []
        for row in self.rows:
            row_12345_id = str(row[self.COL_12345_ID - 1]).strip()
            if row_12345_id != "":
                if include_duplicates == False and row_12345_id in ids:
                    continue
                ids.append(row_12345_id)
        return ids

    
    def get_12345_ids_histogram(self, recurrent_id_only=False):
        ids = self.get_all_12345_ids(include_duplicates=True)
        ids_histogram = {}
        for _id in ids:
            ids_histogram[_id] = ids_histogram.get(_id, 0) + 1
        
        if recurrent_id_only:
            keys = tuple(ids_histogram.keys())
            for _id in keys:
                if ids_histogram[_id] < 2:
                    ids_histogram.pop(_id)
        
        return ids_histogram


    def get_recurrent_rows(self):
        recurrent_ids = tuple(self.get_12345_ids_histogram(recurrent_id_only=True).keys())
        recurrent_rows = []
        for row in self.rows:
            if row[self.COL_12345_ID - 1] in recurrent_ids:
                recurrent_rows.append(row)
        return recurrent_rows

    
    def get_earliest_row(self):
        earliest_row = None
        for row in self.rows:
            if earliest_row is None:
                earliest_row = row
            elif str_to_dt(row[self.COL_RECEIVE_TIME - 1].strip()) < str_to_dt(earliest_row[self.COL_RECEIVE_TIME - 1].strip()):
                earliest_row = row
        return earliest_row


    def get_latest_row(self):
        return [0, 1]


class SangaoTemplate(object):
    def __init__(self, rows):
        self.rows = rows
        template_file = os.path.join(os.path.dirname(__file__), "templates", "sangao_template.xlsx")
        self.template_workbook = openpyxl.load_workbook(template_file)
        self.sheet_1 = self.template_workbook["Sheet1"]
        self.sheet_2 = self.template_workbook["Sheet2"]

        self.write_rows_to_template(self.rows)


    def write_rows_to_template(self, rows):
        COL_DIFFERENCE = 6
        row_count = len(rows)
        for row_index in range(0, row_count):
            col_index = 1 + COL_DIFFERENCE
            for cell_value in rows[row_index]:
                self.sheet_1.cell(row=row_index+2, column=col_index).value = cell_value
                col_index += 1

    
    def fix_validation(self, sheet):
        validator = openpyxl.worksheet.datavalidation.DataValidation(type="list", formula1="=Sheet2!$A$2:$A$118", allow_blank=True)
        validator.add("A2:A1048576")
        sheet.add_data_validation(validator)


    def save(self, save_path):
        self.template_workbook.save(save_path)


class ValidationReport(object):
    def __init__(self, zhengfubiao, sangaobiao):
        self.zhengfubiao = zhengfubiao
        self.sangaobiao = sangaobiao
        self.has_missing_ids = False
        self.missing_ids = []
        self.recurrent_id_histogram = self.sangaobiao.get_12345_ids_histogram(recurrent_id_only=True)
        self.has_recurrent_rows = len(self.recurrent_id_histogram) > 0

        zhengfubiao_ids = zhengfubiao.get_all_ids(include_duplicates=False)
        sangaobiao_ids = sangaobiao.get_all_12345_ids(include_duplicates=True)

        for zhengfubiao_id in zhengfubiao_ids:
            if zhengfubiao_id not in sangaobiao_ids:
                self.missing_ids.append(zhengfubiao_id)
        
        if self.missing_ids != []:
            self.has_missing_ids = True


    def generate_report_text(self):
        jinja2_env = Environment(
            loader=PackageLoader('xbrx_12345_excel_tool', 'templates'),
            autoescape=select_autoescape(['html', 'xml'])
        )
        template = jinja2_env.get_template("validation_report.html")

        recurrent_rows = self.sangaobiao.get_recurrent_rows()
        recurrent_rows.sort(key=lambda row: row[1]) # combine rows with same 12345 id

        ctx = {
            "validation_datetime": get_now_report_str(),
            "zhengfubiao": self.zhengfubiao,
            "sangaobiao": self.sangaobiao,

            "has_missing_ids": self.has_missing_ids,
            "missing_ids": self.missing_ids,
            "missing_rows": self.zhengfubiao.get_rows_by_ids(self.missing_ids),

            "recurrent_rows": recurrent_rows,
            "recurrent_id_histogram": self.recurrent_id_histogram,
            "recurrent_rows_count": len(self.recurrent_id_histogram),
            "has_recurrent_rows": self.has_recurrent_rows,
        }

        report_text = template.render(ctx=ctx)
        return report_text


    def save(self):
        now = datetime.datetime.now()
        filename = f"复核记录_{dt_to_datetime_str(now)}.html"
        report_text = self.generate_report_text()
        validation_logs_dir = make_config_dirs()[2]
        file_path = os.path.join(validation_logs_dir, filename)
        with open(file_path, "w", encoding="utf-8") as f:
            f.write(report_text)
        return validation_logs_dir
        
